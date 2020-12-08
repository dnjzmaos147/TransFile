from openpyxl import Workbook
import re
import json

import os

from openpyxl import load_workbook




fname = '개체명가이드'
angelEx=load_workbook(filename='./{}.xlsx'.format(fname))

sheet = angelEx['제외단어']


output_json = []
        
drop_word = []


multiple_cells = sheet['B1':'B281']
for row in multiple_cells:
    

    for cell in row:

        drop_word.append(cell.value)
   
    


 
path_dir = 'C:/Users/82109/Desktop/2cha/labeljs'
 
file_list = os.listdir(path_dir)

data = []
# 워크북 생성
wb = Workbook()
 
# 워크북 활성화
ws = wb.active

with open('new_plz.json', 'rt', encoding='UTF8') as json_file:
    cate_dict = json.load(json_file)


for flist in file_list:
    fileName = flist


    # fileName = fileName.replace('.xlsx', '')

    with open('./labeljs/{}'.format(fileName), 'rt', encoding='UTF8') as json_file:
        js_data = json.load(json_file)


    def catch_drop(text):
        for word in drop_word:
            if text == word:
                return False
        return True


    def match_cate(js_obj):
        for ca in cate_dict:
            if js_obj['value'] == ca['value']:
                return ca
        return js_obj


    def intent_duplicate(obj):
        for ob in data:
            if ob['value'] == obj['value']:
                return False
        return True

    jsdata = {}
    for line in js_data:
        jsdata = js_data[line]
        break




    for li in jsdata:
        for onejs in jsdata[li]:
            if catch_drop(onejs['value']):

                if intent_duplicate(onejs):
                    match_result = match_cate(onejs)
                    data.append(match_result)


for wo in data:
    if wo['standard'] == 'TTA Basic':
        wo['standard'] = 'TTA_BASIC'
    ws.append([
        wo['standard'],
        wo['categoryID'],
        wo['type'],
        wo['value'],
    ])



wb.save('dict.xlsx')

